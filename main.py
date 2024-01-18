# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import tkinter as tk
from tkinter import filedialog
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Frame, Paragraph, KeepInFrame
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import utils
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import traceback
import csv
import os

fileName = ""
saveLoc = ''
sheetNo = 0
#sImg = []
def readEx(sh):
    wb = load_workbook(sh)
    wbs = wb.worksheets[sheetNo]
    #wbx = wbs['A2':'D{0}'.format(countRow(wbs))]
    #extractImg(wbs)
    createPrintPDF(wbs)
    return(wbs)

def extractImg(wb):
    imgLoad = SheetImageLoader(wb)
    #global sImg
    sImg = []
    for i in range(1,countRow(wb)):
        if imgLoad.image_in('A{0}'.format(i)):
            sImg.append(imgLoad.get('A{0}'.format(i)))
        else:
            sImg.append(None)
    #print(sImg)
    return sImg

def countRow(ws):
    rc = 0
    for row in ws:
        if any(cell.value is not None for cell in row):
            rc+=1
    return rc
    

def createPrintPDF(wbk):
    pageCount = 0
    imgx = 1
    imgy = 24.95
    titx = 1.83 #against image L?
    tity = 24.95+1.3-1.74
    subx = 1.83
    suby = 24.95+1.3-1.74
    qrcx = 16.2
    qrcy = 24.95

    pdfmetrics.registerFont(TTFont('HelveticaNeue', 'HelveticaNeue-01.ttf'))
    pdfmetrics.registerFont(TTFont('HelveticaNeue-CondensedBold', 'HelveticaNeue-CondensedBold-05.ttf'))

    imgs = extractImg(wbk)
    #print(imgs)
    #for total rows: if modulo 6=0 new page
    c = canvas.Canvas('{0}{1}labels.pdf'.format(saveLoc,'/'),pagesize=A4)
    for crow in range(2,countRow(wbk)+2):
        if ((crow-2)%6==0 or crow==countRow(wbk)+1):
            if (crow!=2):
                c.showPage()
                if (crow==countRow(wbk)+1):
                    c.save()
                    
                    
                imgx = 1
                imgy = 24.95
                titx = 1.83 #against image L?
                tity = 24.95+1.3-1.74
                subx = 1.83
                suby = 24.95+1.3-1.74
                qrcx = 16.2
                qrcy = 24.95
                if (crow >= countRow(wbk)+1):
                    break;         
            pageCount+=1
        #img
        #fImg = Frame(imgx*cm,imgy*cm,3.8*cm,3.8*cm,showBoundary=1)
        try:
            #print(scaleImage(imgs[crow-2]))
            #c.drawImage(utils.ImageReader(imgs[crow-2]),x=imgx*cm, y=imgy*cm,width=3.8*cm,height=3.8*cm)
            pass
        except:
            pass
        #fImg.addFromList(dimg, c)
        #fImg.drawBoundary(c)
        imgy-=4.8
        
        #title
        fTit = Frame(titx*cm, tity*cm, 11.4*cm, 2.5*cm, showBoundary=1)
        titStyle = ParagraphStyle('title', fontName='HelveticaNeue-CondensedBold', fontSize=24, alignment=0, wordWrap=None, leading=75)
        if wbk.cell(row=crow,column=2).value == None:
            dtittemp = ''
        else:
            dtittemp = [Paragraph(str(wbk.cell(row=crow,column=2).value).upper(),titStyle)]
        dtit = KeepInFrame(11.4*cm, 2*cm, dtittemp, mode='overflow', vAlign='TOP', hAlign='LEFT', fakeWidth=False)
        fTit.addFromList([dtit], c)
        #fTit.drawBoundary(c)
        tity-=4.8
        
        
        #subtitle
        #fSub = Frame(subx*cm,suby*cm, 11.4*cm,1.3*cm,showBoundary=1)
        #subStyle = ParagraphStyle('subtitle', fontName='HelveticaNeue', fontSize=14, alignment=0, wordWrap=None, leading=75)
        #if wbk.cell(row=crow,column=3).value == None:
        #    dsubtemp = ''
        #else:
        #    dsubtemp = [Paragraph(str(wbk.cell(row=crow,column=3).value).upper(),subStyle)]
        #dsub = KeepInFrame(11.4*cm, 1.3*cm, dsubtemp, mode='truncate', vAlign='TOP', hAlign='LEFT', fakeWidth=False)
        #fSub.addFromList([dsub], c)
        #fSub.drawBoundary(c)

        subStyle = ParagraphStyle('subtitle', fontName='HelveticaNeue', fontSize=14, alignment=0, wordWrap=None, leading=75)
        p = Paragraph("TEST", style=subStyle)
        p.wrapOn(c, c.width, c.height)
        p.drawOn(c, subx*cm, suby*cm)
        suby-=4.8
        
        #qrcode
        fQrc = Frame(qrcx*cm,qrcy*cm,3.8*cm,3.8*cm,showBoundary=1)
        #fQrc.drawBoundary(c)
        qrcy-=4.8

def loadGUI():
    window = tk.Tk()
    window.geometry("600x115")
    window.resizable(False,False)
    window.title("Kevin's Super Duper Ultimate Label Printer Series 5000")
    
    def fileUp(event = None):
        fn = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx")])
        global fileName 
        fileName = fn
        fteT.set(fileName)
        fte["text"] = fteT.get()
        
    def saveUp(event = None):
        sl = filedialog.askdirectory()
        global saveLoc
        saveLoc = sl
        steT.set(saveLoc)
        dle["text"] = steT.get()
    
    def preRunCheck(event = None):
        if (fteT.get() == '' or steT.get()==''):
            el["text"]="After a not-so-comprehensive internal investigation we've determined this incident to be user error"
        else:
            try:
                mS = readEx(fileName)
                #for c1,c2,c3,c4 in mS:
                #    print("{0} {1} {2} {3}".format(c1.value, c2.value, c3.value, c4.value))
            except Exception as error:
                print(error)
                el["text"]= error,str(traceback.extract_stack()[-1][1])
    
    ft = tk.Label(text="xlsx File:")
    ft.place(width=50,height=25,x=10,y=10)
    
    fteT = tk.StringVar(window, '')
    fte = tk.Label(window, text=fteT.get())
    fte.place(width=400,height=25,x=70,y=10)
    
    fbut = tk.Button(window, text="Open", command=fileUp)
    fbut.place(width=50,height=25,x=480,y=10)
    
    rbut = tk.Button(window, text="Run", command=preRunCheck)
    rbut.place(width=50,height=25,x=540,y=10)
    
    dl = tk.Label(text="Save to:")
    dl.place(width=50,height=25,x=10,y=45)
    
    steT = tk.StringVar(window, '')
    dle = tk.Label(window,text=steT.get())
    dle.place(width=400,height=25,x=70,y=45)
    
    dbut = tk.Button(window,text="Open", command=saveUp)
    dbut.place(width=50,height=25,x=480,y=45)
    
    el = tk.Label(window,text='',wraplength=440)
    el.place(width=440,height=25,x=10,y=80)
    
    window.mainloop()

if __name__ == '__main__':
    loadGUI()